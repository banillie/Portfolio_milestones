''' This programme calculates time difference between reported milestones for each project

input documents:
1) Three quarters master information, typically: latest quarter data, last quarter data, year ago quarter data
2) user needs to set date of interest

output document:
1) individual excel workbooks with project milestone analysis
2) individual excel workbooks highlight where reported project milestones have changed.

See instructions below on how to operate.

'''

import datetime
from bcompiler.utils import project_data_from_master
from openpyxl import Workbook
from openpyxl.styles import Font
from milestone_functions import all_milestone_data_bulk, ap_p_milestone_data_bulk, assurance_milestone_data_bulk

'''Function to filter out ALL milestone data'''
def all_milestone_data_bulk(project_list, master_data):
    upper_dict = {}

    for name in project_list:
        try:
            p_data = master_data[name]
            lower_dict = {}
            for i in range(1, 50):
                try:
                    try:
                        lower_dict[p_data['Approval MM' + str(i)]] = \
                            {p_data['Approval MM' + str(i) + ' Forecast / Actual']: p_data[
                                'Approval MM' + str(i) + ' Notes']}
                    except KeyError:
                        lower_dict[p_data['Approval MM' + str(i)]] = \
                            {p_data['Approval MM' + str(i) + ' Forecast - Actual']: p_data[
                                'Approval MM' + str(i) + ' Notes']}

                    lower_dict[p_data['Assurance MM' + str(i)]] = \
                        {p_data['Assurance MM' + str(i) + ' Forecast - Actual']: p_data[
                                'Assurance MM' + str(i) + ' Notes']}
                except KeyError:
                    pass

            for i in range(18, 67):
                try:
                    lower_dict[p_data['Project MM' + str(i)]] = \
                        {p_data['Project MM' + str(i) + ' Forecast - Actual']: p_data['Project MM' + str(i) + ' Notes']}
                except KeyError:
                    pass
        except KeyError:
            lower_dict = {}

        upper_dict[name] = lower_dict

    return upper_dict

'''Function to filter out approval and project delivery milestones'''
def ap_p_milestone_data_bulk(project_list, master_data):
    upper_dict = {}

    for name in project_list:
        try:
            p_data = master_data[name]
            lower_dict = {}
            for i in range(1, 50):
                try:
                    try:
                        lower_dict[p_data['Approval MM' + str(i)]] = \
                            {p_data['Approval MM' + str(i) + ' Forecast / Actual'] : p_data['Approval MM' + str(i) + ' Notes']}
                    except KeyError:
                        lower_dict[p_data['Approval MM' + str(i)]] = \
                            {p_data['Approval MM' + str(i) + ' Forecast - Actual'] : p_data['Approval MM' + str(i) + ' Notes']}

                except KeyError:
                    pass

            for i in range(18, 67):
                try:
                    lower_dict[p_data['Project MM' + str(i)]] = \
                        {p_data['Project MM' + str(i) + ' Forecast - Actual'] : p_data['Project MM' + str(i) + ' Notes']}
                except KeyError:
                    pass
        except KeyError:
            lower_dict = {}

        upper_dict[name] = lower_dict

    return upper_dict

'''Function to filter out assurance milestone data'''
def assurance_milestone_data_bulk(project_list, master_data):
    upper_dict = {}

    for name in project_list:
        try:
            p_data = master_data[name]
            lower_dict = {}
            for i in range(1, 50):
                lower_dict[p_data['Assurance MM' + str(i)]] = \
                    {p_data['Assurance MM' + str(i) + ' Forecast - Actual']: p_data['Assurance MM' + str(i) + ' Notes']}

            upper_dict[name] = lower_dict
        except KeyError:
            upper_dict[name] = {}

    return upper_dict

'''Function calculates the time difference of reported milestone between two quarters'''
def project_time_difference(proj_m_data_1, proj_m_data_2):
    upper_dict = {}

    for proj_name in proj_m_data_1:
        td_dict = {}
        for milestone in proj_m_data_1[proj_name]:
            if milestone is not None:
                milestone_date = tuple(proj_m_data_1[proj_name][milestone])[0]
                try:
                    if date_of_interest <= milestone_date:
                        try:
                            old_milestone_date = tuple(proj_m_data_2[proj_name][milestone])[0]
                            time_delta = (milestone_date - old_milestone_date).days  # time_delta calculated here
                            if time_delta == 0:
                                td_dict[milestone] = 0
                            else:
                                td_dict[milestone] = time_delta
                        except (KeyError, TypeError):
                            td_dict[milestone] = 'not reported' # not reported that quarter
                except (KeyError, TypeError):
                    td_dict[milestone] = 'No date provided' # date has now been removed

        upper_dict[proj_name] = td_dict

    return upper_dict

def filter_group(dictionary, group_of_interest):
    project_list = []
    for project in dictionary:
        if dictionary[project]['DfT Group'] == group_of_interest:
            project_list.append(project)

    return project_list

def filter_gmpp(dictionary):
    project_list = []
    for project in dictionary:
        if dictionary[project]['GMPP - IPA ID Number'] is not None:
            project_list.append(project)

    return project_list

def put_into_wb_all(name, t_dict, td_dict, td_dict2):
    wb = Workbook()
    ws = wb.active

    row_num = 2
    for i, milestone in enumerate(td_dict[name].keys()):
        ws.cell(row=row_num + i, column=1).value = name
        ws.cell(row=row_num + i, column=2).value = milestone
        try:
            milestone_date = tuple(t_dict[name][milestone])[0]
            ws.cell(row=row_num + i, column=3).value = milestone_date
        except KeyError:
            ws.cell(row=row_num + i, column=3).value = 0

        try:
            value = td_dict[name][milestone]
            try:
                if int(value) > 0:
                    ws.cell(row=row_num + i, column=4).value = '+' + str(value) + ' (days)'
                elif int(value) < 0:
                    ws.cell(row=row_num + i, column=4).value = str(value) + ' (days)'
                elif int(value) == 0:
                    ws.cell(row=row_num + i, column=4).value = value
            except ValueError:
                ws.cell(row=row_num + i, column=4).value = value
        except KeyError:
            ws.cell(row=row_num + i, column=4).value = 0

        try:
            value = td_dict2[name][milestone]
            try:
                if int(value) > 0:
                    ws.cell(row=row_num + i, column=5).value = '+' + str(value) + ' (days)'
                elif int(value) < 0:
                    ws.cell(row=row_num + i, column=5).value = str(value) + ' (days)'
                elif int(value) == 0:
                    ws.cell(row=row_num + i, column=5).value = value
            except ValueError:
                ws.cell(row=row_num + i, column=5).value = value
        except KeyError:
            ws.cell(row=row_num + i, column=5).value = 0

        try:
            milestone_date = tuple(t_dict[name][milestone])[0]
            ws.cell(row=row_num + i, column=6).value = t_dict[name][milestone][milestone_date] # provided notes
        except IndexError:
            ws.cell(row=row_num + i, column=6).value = 0

    ws.cell(row=1, column=1).value = 'Project'
    ws.cell(row=1, column=2).value = 'Milestone'
    ws.cell(row=1, column=3).value = 'Date'
    ws.cell(row=1, column=4).value = '3/m change (days)'
    ws.cell(row=1, column=5).value = '1/y change (days)'
    ws.cell(row=1, column=6).value = 'Notes'


    return wb

'''Function that checks whether reported milestone keys have changed between quarters'''
def check_m_keys_in_excel(name, t_dict_one, t_dict_two, t_dict_three):
    wb = Workbook()
    ws = wb.active
    red_text = Font(color="00fc2525")

    row_num = 2

    one = list(t_dict_one[name].keys())
    [x for x in one if x is not None].sort()
    two = list(t_dict_two[name].keys())
    [x for x in two if x is not None].sort()
    three = list(t_dict_three[name].keys())
    [x for x in three if x is not None].sort()

    long = longest_list(one, two, three)
    for i in range(0, len(long)):
        ws.cell(row=row_num + i, column=1).value = name
        try:
            ws.cell(row=row_num + i, column=2).value = one[i]
        except IndexError:
            pass
        try:
            ws.cell(row=row_num + i, column=3).value = two[i]
            if two[i] not in one:
                ws.cell(row=row_num + i, column=3).font = red_text
        except IndexError:
            pass
        try:
            ws.cell(row=row_num + i, column=4).value = three[i]
            if three[i] not in one:
                ws.cell(row=row_num + i, column=4).font = red_text
        except IndexError:
            pass

    #row_num = row_num + len(long)

    return wb

'''helper function for check_m_keys_in_excle'''
def longest_list(one, two, three):
    list_list = [one, two, three]
    a = len(one)
    b = len(two)
    c = len(three)

    out = [a,b,c]
    out.sort()
    for x in list_list:
        if out[-1] == len(x):
            return x

'''INSTRUCTIONS FOR RUNNING THE PROGRAMME'''

'''1) specify file paths to master data for analysis'''
current_Q_dict = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\'
                                          'core data\\Hs2_NPR_Q1_1918_draft.xlsx')
last_Q_dict = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\'
                                       'core data\\master_4_2018.xlsx')
yearago_Q_dict = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\'
                                          'core data\\master_1_2018.xlsx')

'''2) choose list of projects that require output documents'''

'''all projects in portfolio'''
current_Q_list = list(current_Q_dict.keys())

'''projects by group - in development dont use'''
# group_names = ['Rail Group', 'HSMRPG', 'International Security and Environment', 'Roads Devolution & Motoring']

'''single project'''
#current_Q_list = ['Lower Thames Crossing']

'''3) Specify date after which project milestones should be returned. NOTE: Python date format is (YYYY,MM,DD). So 
far this is normally 6 months prior to the quarter closing'''
date_of_interest = datetime.date(2019, 1, 1)

'''4) enter relevant variables in the below functions. 
The type of milestone you wish to analyse can be specified through choosing
all_milestone_data_bulk, ap_p_milestone_data_bulk, or assurance_milestone_data_bulk functions.
the variables to be placed into the functions are firstly list of projects to be analysed and secondly the quarters
variable above corresponding to the variable you are creating'''
current_milestones_dict = ap_p_milestone_data_bulk(current_Q_list, current_Q_dict)
last_milestones_dict = ap_p_milestone_data_bulk(current_Q_list, last_Q_dict)
oldest_milestones_dict = ap_p_milestone_data_bulk(current_Q_list, yearago_Q_dict)

first_diff_dict = project_time_difference(current_milestones_dict, last_milestones_dict)
second_diff_dict = project_time_difference(current_milestones_dict, oldest_milestones_dict)

'''5) running the milestone comparision programme. you must firstly ensure in the first line of the code that the list 
of projects you want to analysis is consistent with the list you have specified above. you then select the relevant 
variables for the put_into_wb_all function. Finally enter the file path to where output documents should be saved. 
Note keep {} in file name as this is where the project name is recorded in the file title'''
for name in current_Q_list:
    print('Doing milestone movement analysis for ' + str(name))
    wb = put_into_wb_all(name, current_milestones_dict, first_diff_dict, second_diff_dict)
    wb.save('C:\\Users\\Standalone\\Will\\Q1_1920_{}_milestone_movement_analysis.xlsx'.format(name))

'''6) running the milestone key comparision programme. you must firstly ensure in the first line of the code that the list 
of projects you want to analysis is consistent with the list you have specified above. in this case it is simple the three
quarters variable data above, from newest to oldest. Finally enter the file path to where output documents should be saved. 
Note keep {} in file name as this is where the project name is recorded in the file title'''
for name in current_Q_list:
    print('Doing milestone key name checking for ' + str(name))
    wb = check_m_keys_in_excel(name, current_milestones_dict, last_milestones_dict, oldest_milestones_dict)
    wb.save('C:\\Users\\Standalone\\Will\\Q1_1920_{}_milestone_keys_check.xlsx'.format(name))