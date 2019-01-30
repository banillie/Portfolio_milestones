import datetime
from bcompiler.utils import project_data_from_master
from openpyxl import Workbook
#from docx import Document
#from docx.shared import Cm, Inches, Pt

'''The below lists caputure the datamap sections of the master spreadsheet which returm milestone keys and milestone dates
The lists can be altered to look at different milestones'''

'''All milestones'''
milestone_keys = ['Approval MM1', 'Approval MM2', 'Approval MM3', 'Approval MM4', 'Approval MM5', 'Approval MM6', 'Approval MM7', 'Approval MM8', 'Approval MM9', 'Approval MM10', 'Approval MM11', 'Approval MM12', 'Approval MM13', 'Approval MM14', 'Approval MM15', 'Approval MM16', 'Project MM18', 'Project MM19', 'Project MM20', 'Project MM21', 'Project MM22', 'Project MM23', 'Project MM24', 'Project MM25', 'Project MM26', 'Project MM27', 'Project MM28', 'Project MM29', 'Project MM30', 'Project MM31', 'Project MM32', 'Assurance MM1', 'Assurance MM2', 'Assurance MM3', 'Assurance MM4', 'Assurance MM5', 'Assurance MM6', 'Assurance MM7', 'Assurance MM8', 'Assurance MM9', 'Assurance MM10', 'Assurance MM11', 'Assurance MM12', 'Assurance MM13', 'Assurance MM14', 'Assurance MM15', 'Assurance MM16', 'Assurance MM17', 'Assurance MM18']
milestone_dates = ['Approval MM1 Forecast / Actual', 'Approval MM2 Forecast / Actual', 'Approval MM3 Forecast / Actual', 'Approval MM4 Forecast / Actual', 'Approval MM5 Forecast / Actual', 'Approval MM6 Forecast / Actual', 'Approval MM7 Forecast / Actual', 'Approval MM8 Forecast / Actual', 'Approval MM9 Forecast / Actual', 'Approval MM10 Forecast / Actual', 'Approval MM11 Forecast / Actual', 'Approval MM12 Forecast / Actual', 'Approval MM13 Forecast - Actual', 'Approval MM14 Forecast - Actual', 'Approval MM15 Forecast - Actual', 'Approval MM16 Forecast - Actual', 'Project MM18 Forecast - Actual', 'Project MM19 Forecast - Actual', 'Project MM20 Forecast - Actual', 'Project MM21 Forecast - Actual', 'Project MM22 Forecast - Actual', 'Project MM23 Forecast - Actual', 'Project MM24 Forecast - Actual', 'Project MM25 Forecast - Actual', 'Project MM26 Forecast - Actual', 'Project MM27 Forecast - Actual', 'Project MM28 Forecast - Actual', 'Project MM29 Forecast - Actual', 'Project MM30 Forecast - Actual', 'Project MM31 Forecast - Actual', 'Project MM32 Forecast - Actual','Assurance MM1 Forecast - Actual', 'Assurance MM2 Forecast - Actual', 'Assurance MM3 Forecast - Actual', 'Assurance MM4 Forecast - Actual', 'Assurance MM5 Forecast - Actual', 'Assurance MM6 Forecast - Actual', 'Assurance MM7 Forecast - Actual', 'Assurance MM8 Forecast - Actual', 'Assurance MM9 Forecast - Actual', 'Assurance MM10 Forecast - Actual', 'Assurance MM11 Forecast - Actual', 'Assurance MM12 Forecast - Actual', 'Assurance MM13 Forecast - Actual', 'Assurance MM14 Forecast - Actual', 'Assurance MM15 Forecast - Actual', 'Assurance MM16 Forecast - Actual', 'Assurance MM17 Forecast - Actual', 'Assurance MM18 Forecast - Actual']

'''approval and project milestones'''
#milestone_keys = ['Approval MM1', 'Approval MM2', 'Approval MM3', 'Approval MM4', 'Approval MM5', 'Approval MM6',
#                  'Approval MM7', 'Approval MM8', 'Approval MM9', 'Approval MM10', 'Approval MM11', 'Approval MM12',
#                  'Approval MM13', 'Approval MM14', 'Approval MM15', 'Approval MM16', 'Project MM18', 'Project MM19',
#                  'Project MM20', 'Project MM21', 'Project MM22', 'Project MM23', 'Project MM24', 'Project MM25',
#                  'Project MM26', 'Project MM27', 'Project MM28', 'Project MM29', 'Project MM30', 'Project MM31',
#                  'Project MM32']
#milestone_dates = ['Approval MM1 Forecast / Actual', 'Approval MM2 Forecast / Actual', 'Approval MM3 Forecast / Actual',
#                   'Approval MM4 Forecast / Actual', 'Approval MM5 Forecast / Actual', 'Approval MM6 Forecast / Actual',
#                   'Approval MM7 Forecast / Actual', 'Approval MM8 Forecast / Actual', 'Approval MM9 Forecast / Actual',
#                   'Approval MM10 Forecast / Actual', 'Approval MM11 Forecast / Actual',
#                   'Approval MM12 Forecast / Actual', 'Approval MM13 Forecast - Actual',
#                   'Approval MM14 Forecast - Actual', 'Approval MM15 Forecast - Actual',
#                   'Approval MM16 Forecast - Actual', 'Project MM18 Forecast - Actual',
#                   'Project MM19 Forecast - Actual', 'Project MM20 Forecast - Actual', 'Project MM21 Forecast - Actual',
#                   'Project MM22 Forecast - Actual', 'Project MM23 Forecast - Actual', 'Project MM24 Forecast - Actual',
#                   'Project MM25 Forecast - Actual', 'Project MM26 Forecast - Actual', 'Project MM27 Forecast - Actual',
#                   'Project MM28 Forecast - Actual', 'Project MM29 Forecast - Actual', 'Project MM30 Forecast - Actual',
#                   'Project MM31 Forecast - Actual', 'Project MM32 Forecast - Actual']

'''business case approval milestones'''
# milestone_keys = ['Approval MM1', 'Approval MM3', 'Approval MM10']
# milestone_dates = ['Approval MM1 Forecast / Actual', 'Approval MM3 Forecast / Actual', 'Approval MM10 Forecast / Actual']

'''function that filters ALL project reported milestone
into dictionaries for passing into the time_difference function'''

def milestone_dict(name, dictionary, m_list, m_list2):
    m_dict = {}
    if name in dictionary.keys():
        m_dict = {}
        z = dictionary[name]
        # m_list is a list of milestone keys. The below loop puts
        # specific project milestone keys into a list.
        milestone_keys = []
        for x in m_list:
            a = z[x]
            milestone_keys.append(a)

        # m_list2 is a list of milestone dates. The below loop puts
        # specific project milestone dates into a list.
        milestone_dates = []
        for x in m_list2:
            b = z[x]
            milestone_dates.append(b)

        # the below loop places the above lists into dictionary.
        for i in range(0, len(milestone_keys)):
            milestone_key = milestone_keys[i]
            milestone_date = milestone_dates[i]
            # checks for dates being reported as strings. Does not handle
            # error but prints the problemtic data so can be changed in master
            if type(milestone_date) == str:
                print(milestone_key, milestone_date)
            # handles none type reporting for milestone dates. These are not
            # needed in dictionaries.
            if milestone_date != None:
                m_dict[milestone_key] = milestone_date
            else:
                pass

    else:
        m_dict = {}

    # print(m_dict)
    return m_dict


'''Function that filters milestones between the range of interest and
calculates the time delta difference between reported milestone dates'''

def time_difference(project_name, dictionary_1, dictionary_2):
    td_dict = {}
    for milestone in dictionary_1[project_name]:
        milestone_date = dictionary_1[project_name][milestone]
        if date_1 <= milestone_date <= date_2:  # milestone filtered with in selected range
            try:
                old_milestone_date = dictionary_2[project_name][milestone]
                time_delta = (milestone_date - old_milestone_date).days  # time_delta calculated here
                if time_delta == 0:
                    td_dict[milestone] = 0
                else:
                    td_dict[milestone] = time_delta
            except KeyError:
                td_dict[milestone] = 'Not reported'

    # print(td_dict)
    return td_dict


''' function that puts milestones in order and gives them a number '''


def calculate_order(projects, m_dict):
    order_dict = {}
    for x in projects:
        first = m_dict[x]
        first_list = list(first.items())
        filtered_list = []
        for i in first_list:
            if date_1 <= i[1] <= date_2:
                filtered_list.append(i)

        filtered_list.sort(key=lambda x: x[1])

        project_milestone_list = []
        for i in filtered_list:
            project_milestone_list.append(i[0])

        order_dict[x] = project_milestone_list

    return order_dict


def filter_group(dictionary, group_of_interest):
    project_list = []
    for project in dictionary:
        if dictionary[project]['DfT Group'] == group_of_interest:
            project_list.append(project)

    return project_list


'''function places information into a word document. key arguments that are
passed into it are. 
t_dict = time dictionaries structured as (milestone:time). Three in total are passed in.
td_dict = time delta dictioniaries structured as (milestone:time_delta). Two in total
are passed in.
sorted_dict = a milestone dictionary that has been sorted into time order. Structured
as (milestone:time)
doc = word document'''

def put_into_wb(project_list, t_dict, td_dict, td_dict2, wb):
    ws = wb.active

    row_num = 0
    for name in project_list:
        print(name)
        print(t_dict[name].keys())
        for i, milestone in enumerate(t_dict[name].keys()):
            ws.cell(row=row_num+i+2, column=1).value = name
            ws.cell(row=row_num+i+2, column=2).value = milestone
            try:
                ws.cell(row=row_num+i+2, column=3).value = t_dict[name][milestone]
            except KeyError:
                ws.cell(row=row_num + i + 2, column=3).value = 0

            try:
                ws.cell(row=row_num+i+2, column=4).value = td_dict[name][milestone]
            except KeyError:
                ws.cell(row=row_num + i + 2, column=4).value = 0
            try:
                ws.cell(row=row_num+i+2, column=5).value = td_dict2[name][milestone]
            except KeyError:
                ws.cell(row=row_num + i + 2, column=5).value = 0


        row_num = row_num + len(t_dict[name])

    ws.cell(row=1, column=1).value = 'Project'
    ws.cell(row=1, column=2).value = 'Milestone'
    ws.cell(row=1, column=3).value = 'Date'
    ws.cell(row=1, column=4).value = '3/m change'
    ws.cell(row=1, column=5).value = '1/y change'


    print(t_dict)
    print(td_dict)
    print(td_dict2)

    return wb

def printing_three(project_list, t_dict, td_dict, td_dict2, ordered_dict, ws):


    #style = doc.styles['Normal']
    #font = style.font
    #font.name = 'Arial'
    #font.size = Pt(10)

    for name in project_list:
        print(name)
        doc.add_paragraph()

        new_para = doc.add_paragraph()
        sorted_dict = ordered_dict[name]
        # print(sorted_dict)
        heading = str(name)
        new_para.add_run(str(heading)).bold = True
        no_rows = len(sorted_dict) + 1
        table1 = doc.add_table(rows=no_rows, cols=4)
        table1.cell(0, 0).text = 'Milestone'
        table1.cell(0, 1).text = 'Current date'
        table1.cell(0, 2).text = 'Three month change'
        table1.cell(0, 3).text = 'One year change'

        for i, milestone in enumerate(sorted_dict):
            table1.cell(i + 1, 0).width = Cm(8)
            '''structured this way so that milestones are numbered'''
            table1.cell(i + 1, 0).text = str(milestone[0]) + '. ' + str(milestone[1])
            print(milestone)

        '''place  dates into the table'''
        for i, milestone in enumerate(sorted_dict):
            if milestone[1] in t_dict[name]:
                '''date'''
                date = t_dict[name][milestone[1]]
                date = datetime.datetime.strptime(date.isoformat(), '%Y-%M-%d').strftime('%d/%M/%Y')
                table1.cell(i + 1, 1).text = str(date)
                '''time difference from last quarter'''
                td = td_dict[name][milestone[1]]
                table1.cell(i + 1, 2).text = str(td)
                '''time difference from oldest quarter'''
                td_2 = td_dict2[name][milestone[1]]
                table1.cell(i + 1, 3).text = str(td_2)

    return doc


'''Function that runs the programme. This is also where the filtering of dates happens. date filters
are currently set at the global level. This is the standard one. It takes all milestones reporting in
current quarter and prints comparisions against those. It does not check to see whether milestones
have been removed'''


def run_standard_comparator(proj_list, dict_1, dict_2, dict_3):
    #doc = Document()
    wb = Workbook()
    ws = wb.active

    '''gather mini-dictionaries for each quarter'''
    current_milestones_dict = {}
    last_milestones_dict = {}
    oldest_milestones_dict = {}
    for x in proj_list:
        current_milestones = milestone_dict(x, dict_1, milestone_keys, milestone_dates)
        current_milestones_dict[x] = current_milestones
        last_milestones = milestone_dict(x, dict_2, milestone_keys, milestone_dates)
        last_milestones_dict[x] = last_milestones
        oldest_milestones = milestone_dict(x, dict_3, milestone_keys, milestone_dates)
        oldest_milestones_dict[x] = oldest_milestones



    '''calculate time current and last quarter'''
    first_diff_dict = {}
    second_diff_dict = {}
    for x in proj_list:
        first_diff = time_difference(x, current_milestones_dict, last_milestones_dict)
        first_diff_dict[x] = first_diff
        second_diff = time_difference(x, current_milestones_dict, oldest_milestones_dict)
        second_diff_dict[x] = second_diff

    #numbering = calculate_order(proj_list, current_milestones_dict)

    '''This enables datalable numbering to be for entire
    number of milestones being reported. i.e indvidual project numbers do
    not revert to 1. This should make it eaiser to follow data_lable
    indexing on the final chart # NOT USED
    number_dict = {}
    big_list = []
    i = 1
    for project in numbering:
        milestones_list = numbering[project]
        i = i
        for milestone in milestones_list:
            big_list.append((i, milestone))
            i += 1
        number_dict[project] = big_list
        big_list = []'''

    '''place all required information into function'''

    run = put_into_wb(proj_list, current_milestones_dict, first_diff_dict, second_diff_dict, wb)
    #run = printing_three(proj_list, current_milestones_dict, first_diff_dict, second_diff_dict, number_dict, ws)

    return run


'''three quarters master dictionaries are required'''
current_Q_dict = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_3_2018.xlsx')
last_Q_dict = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_2_2018.xlsx')
yearago_Q_dict = project_data_from_master('C:\\Users\\Standalone\\Will\\masters folder\\master_3_2017.xlsx')

'''Pass lists of projects of interest i.e those to include in graph here'''
'''all projects in portfolio'''
current_Q_list = list(current_Q_dict.keys())

'''projects by group'''
# group_names = ['Rail Group', 'HSMRPG', 'International Security and Environment', 'Roads Devolution & Motoring']
# current_Q_list = filter_group(current_Q_dict, 'HSMRPG')
# last_Q_list = filter_group(last_Q_dict, 'HSMRPG')
# new_projects_not_reporting = [x for x in current_Q_list if x not in last_Q_list]
# current_Q_list = sorted([x for x in current_Q_list if x not in new_projects_not_reporting])

'''groups of projects'''
# current_Q_list = ['High Speed Rail Programme (HS2)', 'HS2 Phase 2b', 'HS2 Phase1', 'HS2 Phase2a']      #HS2
#current_Q_list = ['A303 Amesbury to Berwick Down']   # Midlands Rail
# current_Q_list = ['Heathrow Expansion','Crossrail Programme', 'Great Western Route Modernisation (GWRM) including electrification',
#                  'Hexagon', 'HS2 Phase1','Western Rail Link to Heathrow']     #Hearthrow Surface Access
# current_Q_list = ['East Midlands Franchise','South Eastern Rail Franchise Competition', 'West Coast Partnership Franchise',
#                 'Cross Country Rail Franchise Competition']       # Rail Franchising
'''single project'''
#current_Q_list = ['Thameslink Programme']

# sets the date of interest at a global level
date_1 = datetime.date(2018, 10, 1)
date_2 = datetime.date(2030, 10, 1)

# runs the functions that put milestones into word doc
print_miles = run_standard_comparator(current_Q_list, current_Q_dict, last_Q_dict, yearago_Q_dict)
print_miles.save('C:\\Users\\Standalone\\Will\\Q3_1819_milestones_for_banke.xlsx')

# test commit