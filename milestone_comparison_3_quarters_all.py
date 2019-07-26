'''This programme to calculate time difference between reported milestones

input documents:
There quarters master information, typically:
1) latest quarter data
2) last quarter data
3) year ago quarter data

output document:
1) excel workbook with all project milestone information

See instructions on how to operate programme below.

'''

#TODO solve problem re filtering in excle when values have + sign in front of the them

import datetime
from bcompiler.utils import project_data_from_master
from openpyxl import Workbook

'''Function to filter out ALL milestone data'''
def all_milestone_data_bulk(project_list, master_data):
    upper_dict = {}

    for name in project_list:
        try:
            p_data = master_data[name]
            lower_dict = {}
            for i in range(1, 50):
                try:
                    try:
                        lower_dict[p_data['Approval MM' + str(i)]] = \
                            {p_data['Approval MM' + str(i) + ' Forecast / Actual']: p_data[
                                'Approval MM' + str(i) + ' Notes']}
                    except KeyError:
                        lower_dict[p_data['Approval MM' + str(i)]] = \
                            {p_data['Approval MM' + str(i) + ' Forecast - Actual']: p_data[
                                'Approval MM' + str(i) + ' Notes']}

                    lower_dict[p_data['Assurance MM' + str(i)]] = \
                        {p_data['Assurance MM' + str(i) + ' Forecast - Actual']: p_data[
                                'Assurance MM' + str(i) + ' Notes']}
                except KeyError:
                    pass

            for i in range(18, 67):
                try:
                    lower_dict[p_data['Project MM' + str(i)]] = \
                        {p_data['Project MM' + str(i) + ' Forecast - Actual']: p_data['Project MM' + str(i) + ' Notes']}
                except KeyError:
                    pass
        except KeyError:
            lower_dict = {}

        upper_dict[name] = lower_dict

    return upper_dict

'''Function to filter out approval and project delivery milestones'''
def ap_p_milestone_data_bulk(project_list, master_data):
    upper_dict = {}

    for name in project_list:
        try:
            p_data = master_data[name]
            lower_dict = {}
            for i in range(1, 50):
                try:
                    try:
                        lower_dict[p_data['Approval MM' + str(i)]] = \
                            {p_data['Approval MM' + str(i) + ' Forecast / Actual'] : p_data['Approval MM' + str(i) + ' Notes']}
                    except KeyError:
                        lower_dict[p_data['Approval MM' + str(i)]] = \
                            {p_data['Approval MM' + str(i) + ' Forecast - Actual'] : p_data['Approval MM' + str(i) + ' Notes']}

                except KeyError:
                    pass

            for i in range(18, 67):
                try:
                    lower_dict[p_data['Project MM' + str(i)]] = \
                        {p_data['Project MM' + str(i) + ' Forecast - Actual'] : p_data['Project MM' + str(i) + ' Notes']}
                except KeyError:
                    pass
        except KeyError:
            lower_dict = {}

        upper_dict[name] = lower_dict

    return upper_dict

'''Function to filter out assurance milestone data'''
def assurance_milestone_data_bulk(project_list, master_data):
    upper_dict = {}

    for name in project_list:
        try:
            p_data = master_data[name]
            lower_dict = {}
            for i in range(1, 50):
                lower_dict[p_data['Assurance MM' + str(i)]] = \
                    {p_data['Assurance MM' + str(i) + ' Forecast - Actual']: p_data['Assurance MM' + str(i) + ' Notes']}

            upper_dict[name] = lower_dict
        except KeyError:
            upper_dict[name] = {}

    return upper_dict

'''Function that calculates time different between milestone dates'''
def project_time_difference(proj_m_data_1, proj_m_data_2, date_of_interest):
    upper_dict = {}

    for proj_name in proj_m_data_1:
        td_dict = {}
        for milestone in proj_m_data_1[proj_name]:
            if milestone is not None:
                milestone_date = tuple(proj_m_data_1[proj_name][milestone])[0]
                try:
                    if date_of_interest <= milestone_date:
                        try:
                            old_milestone_date = tuple(proj_m_data_2[proj_name][milestone])[0]
                            time_delta = (milestone_date - old_milestone_date).days  # time_delta calculated here
                            if time_delta == 0:
                                td_dict[milestone] = 0
                            else:
                                td_dict[milestone] = time_delta
                        except (KeyError, TypeError):
                            td_dict[milestone] = 'Not reported' # not reported that quarter
                except (KeyError, TypeError):
                    td_dict[milestone] = 'No date provided' # date has now been removed

        upper_dict[proj_name] = td_dict

    return upper_dict


''' One of key functions used for calculating which quarter to baseline data from...
Function returns a dictionary structured in the following way project name[('latest quarter info', 'latest bc'), 
('last quarter info', 'last bc'), ('last baseline quarter info', 'last baseline bc'), ('oldest quarter info', 
'oldest bc')] depending on the amount information available in the data. Only the first three key values are returned, 
to ensure consistency (which is helpful later).'''
def bc_ref_stages(proj_list, q_masters_dict_list):

    output_dict = {}

    for name in proj_list:
        #print(name)
        all_list = []      # format [('quarter info': 'bc')] across all masters including project
        bl_list = []        # format ['bc', 'bc'] across all masters. bl_list_2 removes duplicates
        ref_list = []       # format as for all list but only contains the three tuples of interest
        for master in q_masters_dict_list:
            try:
                bc_stage = master[name]['BICC approval point']
                quarter = master[name]['Reporting period (GMPP - Snapshot Date)']
                tuple = (quarter, bc_stage)
                all_list.append(tuple)
            except KeyError:
                pass

        for i in range(0, len(all_list)):
            bl_list.append(all_list[i][1])

        '''below lines of text from stackoverflow. Question, remove duplicates in python list while 
        preserving order'''
        seen = set()
        seen_add = seen.add
        bl_list_2 = [x for x in bl_list if not (x in seen or seen_add(x))]

        ref_list.insert(0, all_list[0])     # puts the latest info into the list first

        try:
            ref_list.insert(1, all_list[1])    # puts that last info into the list
        except IndexError:
            ref_list.insert(1, all_list[0])

        if len(bl_list_2) == 1:                     # puts oldest info into list (as basline if no baseline)
            ref_list.insert(2, all_list[-1])
        else:
            for i in range(0, len(all_list)):      # puts in baseline
                if all_list[i][1] == bl_list[0]:
                    ref_list.insert(2, all_list[i])

        '''there is a hack here i.e. returning only first three in ref_list. There's a bug which I don't fully 
        understand, but this solution is hopefully good enough for now'''
        output_dict[name] = ref_list[0:3]

    return output_dict

'''Another key function used for calcualting which quarter to baseline data from...
Fuction returns a dictionay structured in the following way project_name[n,n,n]. The n (number) values denote where 
the relevant quarter master dictionary is positions in the list of master dictionaries'''
def get_master_baseline_dict(proj_list, q_masters_dict_list, baseline_dict_list):
    output_dict = {}

    for name in proj_list:
        master_q_list = []
        for key in baseline_dict_list[name]:
            for x, master in enumerate(q_masters_dict_list):
                try:
                    quarter = master[name]['Reporting period (GMPP - Snapshot Date)']
                    if quarter == key[0]:
                        master_q_list.append(x)
                except KeyError:
                    pass

        output_dict[name] = master_q_list

    return output_dict



'''function for putting all data into excel for this programme'''
def put_into_wb_all(project_list, t_dict, td_dict, td_dict2, wb):
    ws = wb.active

    row_num = 2
    for name in project_list:
        for i, milestone in enumerate(td_dict[name].keys()):
            ws.cell(row=row_num + i, column=1).value = name
            ws.cell(row=row_num + i, column=2).value = milestone
            try:
                milestone_date = tuple(t_dict[name][milestone])[0]
                ws.cell(row=row_num + i, column=3).value = milestone_date
            except KeyError:
                ws.cell(row=row_num + i, column=3).value = 0

            try:
                value = td_dict[name][milestone]
                # try:
                    # if int(value) > 0:
                    #     ws.cell(row=row_num + i, column=4).value = '+' + str(value) + ' (days)'
                    # elif int(value) < 0:
                    #     ws.cell(row=row_num + i, column=4).value = str(value) + ' (days)'
                    # elif int(value) == 0:
                ws.cell(row=row_num + i, column=4).value = value
                # except ValueError:
                #     ws.cell(row=row_num + i, column=4).value = value
            except KeyError:
                ws.cell(row=row_num + i, column=4).value = 0

            try:
                value = td_dict2[name][milestone]
                # try:
                    # if int(value) > 0:
                    #     ws.cell(row=row_num + i, column=5).value = '+' + str(value) + ' (days)'
                    # elif int(value) < 0:
                    #     ws.cell(row=row_num + i, column=5).value = str(value) + ' (days)'
                    # elif int(value) == 0:
                ws.cell(row=row_num + i, column=5).value = value
                # except ValueError:
                #     ws.cell(row=row_num + i, column=5).value = value
            except KeyError:
                ws.cell(row=row_num + i, column=5).value = 0

            try:
                milestone_date = tuple(t_dict[name][milestone])[0]
                ws.cell(row=row_num + i, column=6).value = t_dict[name][milestone][milestone_date]  # provides notes
            except IndexError:
                ws.cell(row=row_num + i, column=6).value = 0

        row_num = row_num + len(td_dict[name])

    ws.cell(row=1, column=1).value = 'Project'
    ws.cell(row=1, column=2).value = 'Milestone'
    ws.cell(row=1, column=3).value = 'Date'
    ws.cell(row=1, column=4).value = '3/m change (days)'
    ws.cell(row=1, column=5).value = 'Baseline change (days)'
    ws.cell(row=1, column=6).value = 'Notes'

    return wb

'''
Function that runs this programme...
Notes: 1)It does not check to see whether milestones have been removed, 
'''
def run_milestone_comparator(function, proj_list, q_masters_dict_list, date_of_interest):
    wb = Workbook()

    '''firstly business cases of interest are filtered out by bc_ref_stage function'''
    baseline_bc = bc_ref_stages(proj_list, q_masters_dict_list)
    q_masters_list = get_master_baseline_dict(proj_list, q_masters_dict_list, baseline_bc)

    '''gather mini-dictionaries for each quarter'''

    current_milestones_dict = {}
    last_milestones_dict = {}
    oldest_milestones_dict = {}
    for proj_name in proj_list:
        p_current_milestones_dict = function([proj_name], q_masters_dict_list[q_masters_list[proj_name][0]])
        current_milestones_dict.update(p_current_milestones_dict)
        p_last_milestones_dict = function([proj_name], q_masters_dict_list[q_masters_list[proj_name][1]])
        last_milestones_dict.update(p_last_milestones_dict)
        p_oldest_milestones_dict = function([proj_name], q_masters_dict_list[q_masters_list[proj_name][2]])
        oldest_milestones_dict.update(p_oldest_milestones_dict)

    '''calculate time current and last quarter'''
    first_diff_dict = project_time_difference(current_milestones_dict, last_milestones_dict, date_of_interest)
    second_diff_dict = project_time_difference(current_milestones_dict, oldest_milestones_dict, date_of_interest)

    run = put_into_wb_all(proj_list, current_milestones_dict, first_diff_dict, second_diff_dict, wb)

    return run

'''INSTRUCTIONS FOR RUNNING THE PROGRAMME'''

'''1) load all master quarter data files here'''
q1_1920 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_1_2019_wip'
                                   '_(25_7_19).xlsx')
q4_1819 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_4_2018.xlsx')
q3_1819 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_3_2018.xlsx')
q2_1819 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_2_2018.xlsx')
q1_1819 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_1_2018.xlsx')
q4_1718 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_4_2017.xlsx')
q3_1718 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_3_2017.xlsx')
q2_1718 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_2_2017.xlsx')
q1_1718 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_1_2017.xlsx')
q4_1617 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_4_2016.xlsx')
q3_1617 = project_data_from_master('C:\\Users\\Standalone\\general\\masters folder\\core data\\master_3_2016.xlsx')

'''2) Include in the below list, as the variable names, those quarters to include in analysis'''
list_of_dicts_all = [q1_1920 ,q4_1819, q3_1819, q2_1819, q1_1819, q4_1718, q3_1718, q2_1718, q1_1718, q4_1617, q3_1617]
#list_of_dicts_bespoke = [zero, last]

''' 3) set list of projects to be included in output. Still in development'''
'''option one - all projects'''
latest_q_list = list(q1_1920.keys())

'''option two - group of projects... in development'''
group_projects_list = ['Rail Group', 'HSMRPG', 'International Security and Environment', 'Roads Devolution & Motoring']

'''option three - single project'''
one_proj_list = ['Thameslink Programme']

'''4) Specify date after which project milestones should be returned. NOTE: Python date format is (YYYY,MM,DD)'''
start_date = datetime.date(2012, 1, 1)

'''5) choose the type of variables that you would like to place in run_milestone_comparator function, below. 
The type of milestone you wish to analysis can be specified through choosing
all_milestone_data_bulk, ap_p_milestone_data_bulk, or assurance_milestone_data_bulk functions. This choice should be the 
first to be inserted into the below function. After this select the list of the projects on which to perform analysis 
and then the three quarters data that you have put into variables above, in order of newest to oldest.'''
print_miles = \
    run_milestone_comparator(all_milestone_data_bulk, latest_q_list, list_of_dicts_all, start_date)

'''5) specify file path to output document'''
print_miles.save('C:\\Users\\Standalone\\general\\q1_1920_all_milestone_data.xlsx')